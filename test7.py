import os
import sys
import time
import threading
from bs4 import BeautifulSoup
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import *
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

class ProcessingThread(QThread):
    update_progress = pyqtSignal(int)
    update_status = pyqtSignal(str)
    processing_completed = pyqtSignal()

    def __init__(self, file_path, workbook_path, sheet_name, column_name):
        super().__init__()
        self.file_path = file_path
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.column_name = column_name
        self.stopped = False

    def stop(self):
        self.stopped = True

    def run(self):
        self.process_data()

    def process_data(self):
        try:
            self.update_status.emit("Processing...")

            # Initialize Browser
            self.initialize_selenium()

            # Read Excel file into DataFrame
            df = pd.read_excel(self.file_path, sheet_name=self.sheet_name)
            serial_data = df[self.column_name].tolist()

            wb = Workbook()
            wb.remove(wb["Sheet"])
            machine_list_sheet = wb.create_sheet(title="Machine List")
            warranty_list_sheet = wb.create_sheet(title="Warranty Info")
            failed_list_sheet = wb.create_sheet(title="Failed List")

            self.write_headers(machine_list_sheet, "Machine List")
            self.write_headers(warranty_list_sheet, "Warranty Info")
            self.write_headers(failed_list_sheet, "Failed List")

            total_serials = len(serial_data)
            processed_serials = 0

            for serial_number in serial_data:
                if self.stopped:
                    break

                machine_info, warranty_info, fail_reason = self.interact_with_page(serial_number)

                if machine_info:
                    self.write_to_sheet(machine_list_sheet, machine_info)
                if warranty_info:
                    self.write_to_sheet(warranty_list_sheet, warranty_info)
                if fail_reason:
                    self.write_to_sheet(failed_list_sheet, {"serial": serial_number, "fail_reason": fail_reason})

                wb.save(self.workbook_path)  # Save workbook after each serial processed

                processed_serials += 1
                progress_percentage = int((processed_serials / total_serials) * 100)
                self.update_progress.emit(progress_percentage)

            self.update_status.emit("Processing completed.")
            self.processing_completed.emit()  # Signal processing completion

        except Exception as e:
            self.update_status.emit(f"An error occurred during processing: {e}")

        finally:
            self.quit_browser()

    def initialize_selenium(self):
        url = "https://support.hp.com/za-en/check-warranty"
        options = Options()
        options.headless = True  # Run browser in headless mode
        options.add_argument("--headless=new")
        self.driver = webdriver.Chrome(options=options)
        self.driver.get(url)
        self.driver.maximize_window()

    def quit_browser(self):
        if hasattr(self, 'driver'):
            self.driver.quit()

    def interact_with_page(self, serial_number):
        self.remove_clutter()

        if self.wait_loading_screen():
            return None, None, "Timed out"

        input_field = WebDriverWait(self.driver, 120).until(
            EC.presence_of_element_located((By.ID, "inputtextpfinder"))
        )

        self.remove_clutter()
        input_field.clear()
        self.remove_clutter()
        input_field.send_keys(serial_number)

        button_submit = WebDriverWait(self.driver, 120).until(
            EC.element_to_be_clickable((By.ID, "FindMyProduct"))
        )

        self.remove_clutter()
        button_submit.click()
        time.sleep(1.5)

        if self.wait_loading_screen():
            return None, None, "Timed out"

        if self.check_serial_exist():
            return None, None, "Serial unable to match"

        if self.check_requires_prod_num():
            return None, None, "Serial requires prod num"

        if self.network_timeout():
            return None, None, "Timed out"

        try:
            WebDriverWait(self.driver, 120).until(
                EC.presence_of_element_located((By.CLASS_NAME, "info-section"))
            )
            WebDriverWait(self.driver, 120).until(
                EC.presence_of_element_located((By.CLASS_NAME, "product-info-text"))
            )
            WebDriverWait(self.driver, 120).until(
                EC.presence_of_element_located((By.ID, "Support_visitMyProductPage"))
            )
        except:
            return None, None, "Warranty sections appear blank"

        html_content = self.driver.page_source
        machine_info, warranty_info = self.extract_warranty_info(serial_number, html_content)

        link_back = WebDriverWait(self.driver, 120).until(
            EC.element_to_be_clickable((By.ID, "back"))
        )
        self.remove_clutter()
        link_back.click()

        return machine_info, warranty_info, None

    def extract_warranty_info(self, serial_number, html_content):
        soup = BeautifulSoup(html_content, 'html.parser')

        machine_info = {}
        warranty_info = {}

        try:
            product_info_div = soup.find(class_='product-info-text')
            if product_info_div:
                product_name = product_info_div.find('h2').string.strip()
                product_number = product_info_div.find_all('span')[1].text

                page_link_elem = soup.find(id="Support_visitMyProductPage")
                if page_link_elem:
                    page_link = "https://support.hp.com" + page_link_elem['href']

                machine_info = {
                    "serial": serial_number,
                    "product_number": product_number,
                    "product_name": product_name,
                    "page_link": page_link,
                }

        except Exception as e:
            print(f"Exception getting machine info: {e}")

        info_sections = soup.find_all(class_='info-section')
        for section in info_sections:
            warranty_info = {}
            coverage_type = ""
            service_type = ""
            start_date = ""
            end_date = ""
            service_level = ""
            deliverables = ""

            coverage_type_elem = section.find(class_='label', string='Coverage type')
            coverage_type = coverage_type_elem.find_next(class_='text').string.strip() if coverage_type_elem else ""

            service_type_elem = section.find(class_='label', string='Service type')
            service_type = service_type_elem.find_next(class_='text').string.strip() if service_type_elem else ""

            start_date_elem = section.find(class_='label', string='Start date')
            start_date = start_date_elem.find_next(class_='text').string.strip() if start_date_elem else ""

            end_date_elem = section.find(class_='label', string='End date')
            end_date = end_date_elem.find_next(class_='text').string.strip() if end_date_elem else ""

            service_level_elem = section.find(class_='label', string='Service level')
            service_level = ', '.join(p.string.strip() for p in service_level_elem.find_next(class_='text').find_all('p')) if service_level_elem else ""

            deliverables_elem = section.find(class_='label', string='Deliverables')
            deliverables = ', '.join(p.string.strip() for p in deliverables_elem.find_next(class_='text').find_all('p')) if deliverables_elem else ""

            warranty_info = {
                'serial': serial_number,
                'product_number': machine_info.get("product_number", ""),
                "product_name": machine_info.get("product_name", ""),
                'coverage_type': coverage_type,
                'service_type': service_type,
                'start_date': start_date,
                'end_date': end_date,
                'service_level': service_level,
                'deliverables': deliverables,
                'page_link': machine_info.get("page_link", "")
            }

        return machine_info, warranty_info

    def remove_clutter(self):
        try:
            self.driver.execute_script("document.getElementById('onetrust-consent-sdk').style.display = 'none';")
        except:
            pass
        try:
            self.driver.execute_script("document.getElementById('MDigitalLightboxWrapper').style.display = 'none';")
        except:
            pass
        try:
            self.driver.execute_script("document.body.style.overflow = 'auto';")
        except:
            pass

    def wait_loading_screen(self):
        time.sleep(1.5)
        try:
            WebDriverWait(self.driver, 120).until_not(
                EC.presence_of_element_located((By.CLASS_NAME, "loading_screen_wrapper"))
            )
            return False
        except TimeoutException:
            return True

    def check_serial_exist(self):
        try:
            element_locator = (By.CSS_SELECTOR, "p.errorTitle")
            WebDriverWait(self.driver, 1.5).until(
                EC.visibility_of_element_located(element_locator)
            )
            return True
        except:
            return False

    def network_timeout(self):
        try:
            element_locator = (By.ID, "modal-heading")
            WebDriverWait(self.driver, 1.5).until(
                EC.visibility_of_element_located(element_locator)
            )
            return True
        except:
            return False

    def check_requires_prod_num(self):
        try:
            element_locator = (By.CSS_SELECTOR, "p.field-info.errorTxt.is-invalid-field")
            WebDriverWait(self.driver, 1.5).until(
                EC.visibility_of_element_located(element_locator)
            )
            return True
        except:
            return False

    def write_headers(self, sheet, title):
        if title == "Machine List":
            column_headers = ["serial", "product_number", "product_name", "page_link"]
        elif title == "Warranty Info":
            column_headers = [
                "serial",
                "product_number",
                "product_name",
                "coverage_type",
                "service_type",
                "start_date",
                "end_date",
                "service_level",
                "deliverables",
                "page_link"
            ]
        elif title == "Failed List":
            column_headers = ["serial", "fail_reason"]
        sheet.append([column for column in column_headers])

    def write_to_sheet(self, sheet, data):
        sheet.append([data[column] for column in data])


class MyWindow(QWidget):
    processing_finished = pyqtSignal()  # Signal to indicate processing completion

    def __init__(self):
        super().__init__()
        self.setWindowTitle('HP Warranty Check')
        self.setWindowIcon(QIcon('./assets/icon.png'))

        self.general_label = QLabel('HP Warranty Check')
        self.path_text_field = QLineEdit()
        self.path_text_field.setPlaceholderText('Excel File Path')
        self.path_text_field.setReadOnly(True)
        self.upload_btn = QPushButton('Upload')

        self.sheet_name_text_field = QLineEdit()
        self.sheet_name_text_field.setPlaceholderText('Sheet Name')
        self.column_name_text_field = QLineEdit()
        self.column_name_text_field.setPlaceholderText('Column Name')

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)

        self.file_selector_btn = QPushButton('Select Excel File')
        self.file_selector_btn.clicked.connect(self.select_file)

        self.upload_btn.clicked.connect(self.upload_clicked)

        main_layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        input_layout = QHBoxLayout()

        button_layout.addWidget(self.upload_btn)
        input_layout.addWidget(self.sheet_name_text_field)
        input_layout.addWidget(self.column_name_text_field)

        main_layout.addWidget(self.general_label)
        main_layout.addWidget(self.path_text_field)
        main_layout.addWidget(self.file_selector_btn)
        main_layout.addLayout(input_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.progress_bar)

        self.setLayout(main_layout)
        self.adjustSize()
        self.setFixedSize(400, 260)
        self.move(20, 20)
        #self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)

        self.status_label = QLabel()
        main_layout.insertWidget(1, self.status_label)

        # Exit button
        self.exit_button = QPushButton('Exit')
        self.exit_button.clicked.connect(self.exit_application)
        main_layout.addWidget(self.exit_button)

        # Connections for processing thread signals
        self.processing_thread = None
        self.processing_finished.connect(self.processing_completed_handler)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Select Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if file_path:
            self.path_text_field.setText(file_path)

    def upload_clicked(self):
        file_path = self.path_text_field.text().strip('"\'')
        sheet_name = self.sheet_name_text_field.text().strip()
        column_name = self.column_name_text_field.text().strip()

        if not (sheet_name and column_name):
            self.status_label.setText("Please enter sheet and column names.")
            return

        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            if column_name not in df.columns:
                self.status_label.setText("Column not found.")
                return
        except Exception as e:
            self.status_label.setText("Sheet not found.")
            return

        new_file_name, _ = QFileDialog.getSaveFileName(self, 'Select save location and name for excel file: ', '', 'Excel Files (*.xlsx)')

        if new_file_name:
            # Create the workbook with headers
            wb = Workbook()
            machine_list_sheet = wb.create_sheet(title="Machine List")
            warranty_list_sheet = wb.create_sheet(title="Warranty Info")
            failed_list_sheet = wb.create_sheet(title="Failed List")

            self.processing_thread = ProcessingThread(file_path, new_file_name, sheet_name, column_name)
            self.processing_thread.update_progress.connect(self.update_progress_bar)
            self.processing_thread.update_status.connect(self.update_status_label)
            self.processing_thread.processing_completed.connect(self.enable_widgets)
            self.progress_bar.setVisible(True)
            self.upload_btn.setEnabled(False)
            self.file_selector_btn.setEnabled(False)
            self.sheet_name_text_field.setEnabled(False)
            self.column_name_text_field.setEnabled(False)
            self.progress_bar.setValue(0)
            self.processing_thread.start()

    def update_progress_bar(self, percentage):
        self.progress_bar.setValue(percentage)

    def update_status_label(self, message):
        self.status_label.setText(message)

    def exit_application(self):
        # Stop processing thread if running
        if self.processing_thread and self.processing_thread.isRunning():
            self.processing_thread.stop()
            self.processing_thread.quit()

        # Quit Selenium browser if running
        if self.processing_thread and hasattr(self.processing_thread, 'driver'):
            self.processing_thread.quit_browser()

        # Close the application
        QApplication.quit()

    def processing_completed_handler(self):
        # Enable buttons after processing completion
        self.progress_bar.setVisible(False)
        self.enable_widgets()

    def enable_widgets(self):
        self.upload_btn.setEnabled(True)
        self.file_selector_btn.setEnabled(True)
        self.sheet_name_text_field.setEnabled(True)
        self.column_name_text_field.setEnabled(True)


# Create the application instance
app = QApplication(sys.argv)

# Create the main window instance
window = MyWindow()

# Show the main window
window.show()

# Execute the application
sys.exit(app.exec_())
